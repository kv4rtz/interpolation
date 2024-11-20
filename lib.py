from string import ascii_uppercase
import openpyxl.styles

def to_power_symbol(x):
    superscript_digits = {
        0: '⁰', 1: '¹', 2: '²', 3: '³', 4: '⁴', 5: '⁵', 
        6: '⁶', 7: '⁷', 8: '⁸', 9: '⁹'
    }
    return ''.join(superscript_digits[int(digit)] for digit in str(x))

def get_filename_ext(filename: str, ext: str):
    return f'./excel/{filename}.{ext}'

def get_delta(y: list):
    deltas = []
    while len(y) > 1:
        next_deltas = []
        for i in range(len(y) - 1):
            result = y[i + 1] - y[i]
            next_deltas.append(result)
        deltas.append(next_deltas)
        y = next_deltas
    return deltas

def render_lagrange_table(ws, x_table: list[int], y_table: list[int], func):
    ws['A1'] = 'x'
    for i in range(len(x_table)):
        ws[f'A{i + 2}'] = x_table[i]
    ws['B1'] = 'y'
    for i in range(len(y_table)):
        ws[f'B{i + 2}'] = y_table[i]

    for i in ws[f'A1:B1']:
        for j in i:
            j.fill = openpyxl.styles.PatternFill(fill_type='solid', start_color='f8ff7a')
            
    ws[f'A{len(x_table) + 3}'] = "f(x) ="
    ws[f'B{len(x_table) + 3}'] = str(func)

def render_newton_table(ws, x_table: list[int], y_table: list[int], all_deltas: list[list[int]], func):
    ws['A1'] = 'x'
    for i in range(len(x_table)):
        ws[f'A{i + 2}'] = x_table[i]
    ws['B1'] = 'y'
    for i in range(len(y_table)):
        ws[f'B{i + 2}'] = y_table[i]

    for i in range(len(x_table) - 1):
        column = ascii_uppercase[i + 2]
        cell = f"{column}1"
        ws[cell] = f'Δ{to_power_symbol(i + 1)}y'
        for j in range(len(all_deltas[i])):
            row = j + 2
            column = ascii_uppercase[i + 1]
            if (ws[f'{column}{row + 1}'].value and ws[f'{column}{row}'].value):
                ws[f"{ascii_uppercase[i + 2]}{j + 2}"] = f"={column}{row + 1}-{column}{row}"

    for i in ws[f'A1:{ascii_uppercase[len(x_table)]}1']:
        for j in i:
            j.fill = openpyxl.styles.PatternFill(fill_type='solid', start_color='f8ff7a')

    ws[f'A{len(x_table) + 3}'] = "f(x) ="
    ws[f'B{len(x_table) + 3}'] = str(func)

    ws[f'B{len(x_table) + 3}'].fill = openpyxl.styles.PatternFill(fill_type='solid', start_color='58fc84')
    ws[f'A{len(x_table) + 3}'].fill = openpyxl.styles.PatternFill(fill_type='solid', start_color='58fc84')    

def render_vba_function(wb, func):
    f = str(func).replace('**', ' ^ ')
    vba_code = f'''
    Function Polinom(x As Double) As Double
        Polinom = {f}
    End Function
    '''
    
    vba_module = wb.api.VBProject.VBComponents.Add(1)
    vba_module.CodeModule.AddFromString(vba_code)